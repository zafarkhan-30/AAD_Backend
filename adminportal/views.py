from rest_framework import generics
from rest_framework.parsers import MultiPartParser
from django.contrib.auth.models import Group
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from database.models import *
from .serializers import *
from Allauth.serializers import RegisterSerializer
from rest_framework import status
from .permissions import *
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import filters
from rest_framework.pagination import LimitOffsetPagination
from Allauth.serializers import *
from django.db.models import Count, Q, Case, When, IntegerField, F, Value
from django.db.models.functions import Cast, Substr, Coalesce
from django.db.models.expressions import Func
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
# Create your views here.
from .permissions import IsSupervisor
from excel_response import ExcelResponse
from datetime import datetime
from rest_framework.decorators import api_view
from doctorsApp.permissions import IsMO
from adminportal.utils import get_suspected_disease_counts

class CustomPageNumberPagination(PageNumberPagination):
    page_size = 10  # You can adjust this value based on your requirements
    page_size_query_param = 'page_size'
    max_page_size = 100

class PostUserGroupResquest(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = UpdateSerializer
    def post(self, request ,  *args, **kwargs):
        serializer = self.get_serializer(data = request.data )
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status': 'success',
                'message': 'request raise successfully'
            })
        else:
            return Response(serializer.errors)

class GetGroupList(generics.GenericAPIView):
    permission_classes = [IsAuthenticated , IsSupervisor ]
    serializer_class = GroupListSerializer
    def get(self, request):
        group_list = Group.objects.all()
        serializer = self.get_serializer(group_list , many = True).data
        return Response({'group_list': serializer})

class GetGroupRequestList(generics.ListAPIView):
    serializer_class = GetGroupRequestListSerializer
    permission_classes = [IsAuthenticated , IsAdmin | IsMOH| IsViewAdmin]
    pagination_class = LimitOffsetPagination
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)


    def get_queryset(self):
        queryset = self.model.objects.filter( status=False   )

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(user__name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class updateUserGroupRequest(generics.GenericAPIView):
    serializer_class = UpdateGroupRequest
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated , IsAdmin]


    def patch(self, request ,  id ,  *args, **kwargs):
        try:
            instance = UserApprovalRecords.objects.get(id=id)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)
        serializer = self.get_serializer(instance , data = request.data , partial = True )
        if serializer.is_valid():
            serializer.save()
            return Response({"status" : "success" ,
                            "message" : "User group updated successfully"
                    },status=status.HTTP_201_CREATED)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = value[0]
            return Response({'message': error_message,
                            'status' : 'error'}, status=400)

class UserCountsAPI(APIView):
    # permission_classes = [ IsAuthenticated, IsAdmin]
    def get(self, request, *args, **kwargs):
        all = CustomUser.objects.all()
        CHV_ASHA_count = all.filter(groups__name='CHV-ASHA').count()
        MO_count = all.filter(groups__name='mo').count()
        ANM_count = all.filter(groups__name='healthworker').count()
        return Response({
            'CHV_ASHA_count' : CHV_ASHA_count ,
            'MO_count' : MO_count ,
            'ANM_count' : ANM_count
        } , status = 200)

class InsertUsersByadmin(generics.GenericAPIView):
    # permission_classes = [permissions.IsAuthenticated,]
    serializer_class = AddUserSerializer
    # parser_classes = [MultiPartParser]
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor | IsMOH)


    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        # print(serializer)
        try:
            if serializer.is_valid():

                group = Group.objects.get(name=serializer.validated_data.get("group"))

                user = serializer.save(is_active = True)
                customuser = serializer.validated_data
                data = RegisterSerializer(customuser, context=self.get_serializer_context()).data
                user.groups.add(group)
                addSupervisor = CustomUser.objects.filter(id= user.id).update(created_by_id = request.user.id)
                return Response({
                    "status": "success",
                    "message": "Successfully Inserted.",
                    "data": data,
                })
            else:
                key, value = list(serializer.errors.items())[0]
                # print(key , value)
                error_message = key + " ,"  + value[0]
                return Response({
                    "status": "error",
                    "message": error_message,

                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as ex:
            return Response({
                "status": "error",
                "message": "Error in Field " + str(ex),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetDeactivatedUserList(generics.ListAPIView):
    permission_classes = [IsAuthenticated , IsAdmin]
    pagination_class = LimitOffsetPagination
    serializer_class = GetDeactivatedUserListSerializer
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        ward_name = self.kwargs.get('ward_name')
        group = self.kwargs.get('group')

        queryset = self.model.objects.filter( is_active=False  , created_by__groups__name = 'MOH' ,
                                            userSections__healthPost__ward__wardName = ward_name , groups__name = group ).order_by("-created_date")

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) |
                Q(health_Post__healthPostName__icontains=search_terms) |
                Q(userSections__healthPost__healthPostName__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class InsertUsersByMOH(generics.GenericAPIView):
    # permission_classes = [permissions.IsAuthenticated,]
    serializer_class = AddUserByMOHSerializer
    # parser_classes = [MultiPartParser]
    permission_classes = (IsAuthenticated , IsMOH)


    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        try:
            if serializer.is_valid():
                group = Group.objects.get(name=serializer.validated_data.get("group"))

                user = serializer.save()
                customuser = serializer.validated_data
                data = RegisterSerializer(customuser, context=self.get_serializer_context()).data
                user.groups.add(group)

                addSupervisor = CustomUser.objects.filter(id= user.id).update(created_by_id = request.user.id)
                return Response({
                    "status": "success",
                    "message": "Successfully Inserted.",
                    "data": data,
                })
            else:
                key, value = list(serializer.errors.items())[0]
                error_message = value[0]
                return Response({
                    "status": "error",
                    "message": error_message,

                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as ex:
            return Response({
                "status": "error",
                "message": "Error in Field " + str(ex),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UpdateUserDetails(generics.GenericAPIView):
    serializer_class  = UpdateUsersDetailsSerializer
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor | IsMOH)
    # parser_classes = [MultiPartParser]


    def patch(self, request, pk):
        # print(request.data)
        try:
            instance = CustomUser.objects.get(pk=pk)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)

        serializer = self.get_serializer(instance , data = request.data , partial = True )
        # print(serializer)
        if serializer.is_valid():
            if "newpassword" in  serializer.validated_data:
                instance.set_password(serializer.validated_data.get("newpassword"))
                instance.save()
            serializer.save()
            return Response({"status" : "success" ,
                            "message" : "User details updated successfully"
                    },status=status.HTTP_201_CREATED)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = value[0]
            return Response({'message': error_message,
                            'status' : 'error'}, status=400)

class deleteUser(generics.GenericAPIView):
    serializer_class = DeleteUserSerializer
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor)

    def delete(self, request, id , *args, **kwargs):
        try:
            instance = CustomUser.objects.get(pk=id).delete()
            return Response({'status': 'success', 'message' : 'user deleted'}, status= status.HTTP_200_OK)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)

class AdminChangePasswordView(generics.UpdateAPIView):
    """
    An endpoint for changing password.
    """
    serializer_class = ChangePasswordSerializer
    # model = CustomUser
    # permission_classes = (IsAuthenticated)
    def get_object(self, queryset=None):
        id = self.kwargs.get('id')
        try:
            obj = CustomUser.objects.get(id = id)
        except:
            return Response({'status': 'error',
                'message': 'user details not found'
            }, status= status.HTTP_400_BAD_REQUEST)
        return obj

    def update(self , request, *args, **kwargs):
        self.object = self.get_object()
        serializer = self.get_serializer(data=request.data)

        if serializer.is_valid():
            # Check old password
            # set_password also hashes the password that the user will get
            self.object.set_password(serializer.data.get("newpassword"))
            self.object.save()

            # sendOtp.objects.filter(registerUser_id = self.request.user.id).delete()
            response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'Password updated successfully'
            }
            return Response(response)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class userListAPI(generics.ListAPIView):
    pagination_class = LimitOffsetPagination
    serializer_class = CustomUserSerializer
    model = serializer_class.Meta.model
    # permission_classes = (IsAuthenticated , IsAdmin)
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self ):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        group = self.kwargs.get('group')

        ward_name = self.kwargs.get('ward_name')
        # section = self.model.objects.filter(groups__name = group, userSections__id = 239)
        # for i  in section:
        #     print(i.name)
        if group == 'mo':
            queryset = self.model.objects.filter(groups__name = group , dispensary__ward__wardName = ward_name ).order_by("-created_date")
        else:
            # queryset = self.model.objects.filter(groups__name = group , section__healthPost__ward__wardName = ward_name).order_by("-created_date")
            queryset = self.model.objects.filter(groups__name = group , userSections__healthPost__ward__wardName = ward_name).order_by("-created_date").distinct()

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) |
                Q(health_Post__healthPostName__icontains=search_terms) |
                Q(section__healthPost__healthPostName__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset( )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class GetWardWiseSUerList(generics.ListAPIView):
    permission_classes = [IsAuthenticated ,IsMOH]
    pagination_class = LimitOffsetPagination
    serializer_class = CustomUserSerializer
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        group = self.kwargs.get('group')
        # wardName = self.kwargs.get('ward')
        # print(group , wardName)
        ward_id= self.request.user.ward.id
        queryset = self.model.objects.filter(groups__name = group  , userSections__healthPost__ward__id = ward_id).order_by("-created_date").distinct()

        search_terms = self.request.query_params.get('search', None)
        if search_terms:
            queryset = queryset.filter(Q(section__healthPost__ward__wardName__icontains=search_terms)|
                                        Q(userSections__icontains=search_terms) |
                                        Q(phoneNumber__icontains=search_terms) |
                                        Q(health_Post__healthPostName__icontains=search_terms))

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()

        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class DownloadHealthpostwiseUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, id, *args, **kwargs):
        try:
            healthpost = healthPost.objects.get(pk=id)
        except healthPost.DoesNotExist:
            return Response({
                "message":"No Health post exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        healthpost_related_user = familyMembers.objects.filter(area__healthPost=healthpost)
        today = datetime.today().strftime('%d-%m-%Y')
        healthpost_name = healthpost.healthPostName

        if not healthpost_related_user:
            return Response({
                "message":"No data found for healthpost %s"%(healthpost_name),
                "status":"error"
            } , status=400)

        familyMember = healthpost_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', "Address" ,'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  healthpost_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                                family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(healthpost_name+"_data_"+today)
        wb.save(response)
        return response

class DownloadWardwiseUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, id, *args, **kwargs):
        try:
            ward_obj = ward.objects.get(pk=id)
        except ward.DoesNotExist:
            return Response({
                "message":"No Ward exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        # ward_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost__ward=ward_obj)
        ward_related_user = familyMembers.objects.filter(area__healthPost__ward=ward_obj)
        today = datetime.today().strftime('%d-%m-%Y')
        ward_name = ward_obj.wardName

        if not ward_related_user:
            return Response({
                "message":"No data found for ward %s"%(ward_name),
                "status":"error"
            } , status= 400)

        familyMember = ward_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', "Address" ,'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  ward_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                               family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]


            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("Ward_"+ward_name+"_data_"+today)
        wb.save(response)
        return response

class DownloadAllWardUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, *args, **kwargs):

        ward_related_user = familyMembers.objects.all()
        today = datetime.today().strftime('%d-%m-%Y')

        familyMember = ward_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', 'Address', 'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  ward_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address ,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y %I:%M:%S %p'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                                family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("All_Ward_data_"+today)
        wb.save(response)
        return response

class DownloadDispensarywiseUserList(generics.GenericAPIView):

    # permission_classes = [IsAuthenticated , IsMO ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():

                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request ,id, *args, **kwargs):
        try:
            dispensary_obj = dispensary.objects.get(pk=id)
        except dispensary.DoesNotExist:
            return Response({
                "message":"No Dispensary exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        dispensary_related_user = familyMembers.objects.filter(area__dispensary=dispensary_obj)
        today = datetime.today().strftime('%d-%m-%Y')
        dispensary_name = dispensary_obj.dispensaryName

        if not dispensary_related_user:
            return Response({
                "message":"No data found for dispensary %s"%(dispensary_name),
                "status":"error"
            } , status=400)

        familyMember = dispensary_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', 'Address', 'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  dispensary_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address ,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                               family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(dispensary_name+"_data_"+today)
        wb.save(response)
        return response

class MOHDashboardView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated , IsMOH)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def get(self, request ,  *args, **kwargs):

        healthpost_id = request.query_params.get('healthpost_id', None)
        CHV_ASHA_count = self.CustomUser_queryset.filter(userSections__healthPost__ward__id = request.user.ward_id ,groups__name='CHV-ASHA').distinct().count()
        MO_count = self.CustomUser_queryset.filter(dispensary__ward__id = request.user.ward_id, groups__name='mo').count()
        ANM_count = self.CustomUser_queryset.filter(userSections__healthPost__ward__id = request.user.ward_id ,groups__name='healthworker').distinct().count()
        today = timezone.now().date()
        ward_id = request.user.ward_id

        if healthpost_id:

            # Distinct and common queries of survey data
            healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            healthpost_data = healthpost_queryset.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                todays_citizen_count=Count('id', filter=Q(created_date__date=today), distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                vulnerable_70_Years=Count('id', filter=Q(vulnerable_choices__choice='70+ Years'), distinct=True),
                vulnerable_Physically_handicapped=Count('id', filter=Q(vulnerable_choices__choice='Physically Handicapped'), distinct=True),
                vulnerable_completely_paralyzed_or_on_bed=Count('id', filter=Q(vulnerable_choices__choice='Completely Paralyzed or On bed'), distinct=True),
                vulnerable_elderly_and_alone_at_home=Count('id', filter=Q(vulnerable_choices__choice='Elderly and alone at home'), distinct=True),
                vulnerable_any_other_reason=Count('id', filter=Q(vulnerable_choices__choice='Any other reason'), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                partial_survey_count=Count('id', filter=Q(partialSubmit=True), distinct=True),
                total_family_count=Count('id', distinct=True),
                today_family_count=Count('id', filter=Q(created_date__date=today), distinct=True)
            )

            combined_survey_data = {**healthpost_data, **familySurvey_data}

            Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

        else:
            # Distinct and common queries of survey data
            ward_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=ward_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward__id=ward_id)

            # Ward related survey data
            ward_data = ward_queryset.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                todays_citizen_count=Count('id', filter=Q(created_date__date=today), distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                vulnerable_70_Years=Count('id', filter=Q(vulnerable_choices__choice='70+ Years'), distinct=True),
                vulnerable_Physically_handicapped=Count('id', filter=Q(vulnerable_choices__choice='Physically Handicapped'), distinct=True),
                vulnerable_completely_paralyzed_or_on_bed=Count('id', filter=Q(vulnerable_choices__choice='Completely Paralyzed or On bed'), distinct=True),
                vulnerable_elderly_and_alone_at_home=Count('id', filter=Q(vulnerable_choices__choice='Elderly and alone at home'), distinct=True),
                vulnerable_any_other_reason=Count('id', filter=Q(vulnerable_choices__choice='Any other reason'), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                partial_survey_count=Count('id', filter=Q(partialSubmit=True), distinct=True),
                total_family_count=Count('id', distinct=True),
                today_family_count=Count('id', filter=Q(created_date__date=today), distinct=True)
            )

            combined_survey_data = {**ward_data, **familySurvey_data}

            Questionnaire_queryset = ward_queryset.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

        return Response({
            'CHV_ASHA_count' : CHV_ASHA_count,
            'MO_count' : MO_count,
            'ANM_count' : ANM_count,
            'total_count' : combined_survey_data["total_citizen_count"],
            'todays_count' : combined_survey_data["todays_citizen_count"],
            'partial_survey_count' : combined_survey_data["partial_survey_count"],
            'total_family_count' : combined_survey_data["total_family_count"],
            'today_family_count' : combined_survey_data["today_family_count"],
            'total_cbac_count' : combined_survey_data["total_cbac_count"],
            'citizen_above_60' : combined_survey_data["citizen_above_60"],
            'citizen_above_30' : combined_survey_data["citizen_above_30"],
            'TestReportGenerated' : combined_survey_data["TestReportGenerated"],
            'total_LabTestAdded' : combined_survey_data["total_LabTestAdded"],
            'total_AbhaCreated' : combined_survey_data["total_AbhaCreated"],
            "male" : combined_survey_data["male"],
            "female" : combined_survey_data["female"],
            "transgender" : combined_survey_data["transgender"],
            'hypertension' : combined_survey_data["hypertension"],
            **suspected_disease_counts,
            'blood_collected_home' : combined_survey_data["blood_collected_home"],
            'blood_collected_center' : combined_survey_data["blood_collected_center"],
            'denieded_by_mo_count' : combined_survey_data["denied_by_mo_count"],
            'denieded_by_mo_individual' : combined_survey_data["denied_by_mo_individual"],
            'Referral_choice_Referral_to_Mun_Dispensary' : combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
            'Referral_choice_Referral_to_HBT_polyclinic': combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
            'Referral_choice_Referral_to_Peripheral_Hospital': combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
            'Referral_choice_Referral_to_Medical_College': combined_survey_data["Referral_choice_Referral_to_Medical_College"],
            'Referral_choice_Referral_to_Private_facility': combined_survey_data["Referral_choice_Referral_to_Private_facility"],
            'total_vulnerabel' : combined_survey_data["total_vulnerable"],
            'vulnerabel_70_Years' : combined_survey_data["vulnerable_70_Years"],
            'vulnerabel_Physically_handicapped' : combined_survey_data["vulnerable_Physically_handicapped"],
            'vulnerabel_completely_paralyzed_or_on_bed' : combined_survey_data["vulnerable_completely_paralyzed_or_on_bed"],
            'vulnerabel_elderly_and_alone_at_home' : combined_survey_data["vulnerable_elderly_and_alone_at_home"],
            'vulnerabel_any_other_reason' : combined_survey_data["vulnerable_any_other_reason"]}, status= 200)

#need to recheck
class MOHDashboardExportView(generics.GenericAPIView): #Modified

    permission_classes = ( IsAuthenticated , IsMOH )
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def add_headers(self, sheet, *args):
        """This function takes the sheet and add merged headers to that sheet"""
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def get_queryset(self ):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        queryset = familyMembers.objects.all()

        return queryset

    def get(self, request, *args, **kwargs):

        healthpost_id = request.query_params.get('healthpost_id', None)
        wardId = request.query_params.get('wardId', None)
        group_name = request.query_params.get('group_name', None)
        today = datetime.today().strftime('%d-%m-%Y')

        if not group_name:
            return Response({
                    "message":"Please provide group name",
                    "status":"error"
                }, status=400)

        group_name = group_name.strip()

        if group_name == "healthworker":
            data_list = [['ward Name' , 'Health Post Name' , "ANM/Co-ordinator" , "ANM Mobile number" ,'Families Enrolled'  , 'Citizens Enrolled' , 'CBAC Filled' ,
                        'Citizens 60 years + enrolled', 'Citizens 30 years + enrolled' , "Males Enrolled" ,  "Females Enrolled" ,  "Transgender Enrolled",
                        "ABHA ID Generated" , 'Diabetes' , 'Hypertension' ,  'Oral Cancer' , 'Cervical cancer' , 'COPD' , 'Eye Disorder' ,
                        'ENT Disorder' ,  'Asthma' , 'Alzheimers' ,  'TB' , 'leprosy','Breast cancer' , 'Other Communicable' ,
                        'Blood collected at home' , 'blood collected at center' , 'Blood Collection Denied By AMO' ,  'Blood Collection Denied By Citizen',
                        'Total Reports Generated' , 'Tests Assigned' ,
                        'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment' ,
                        'Referral to HBT polyclinic for Physician consultation',
                        'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                        'Referral to Medical College for management of Complication',
                        'Referral to Private facility',
                        'Vulnerable Citizen' ]]
        elif group_name == "CHV-ASHA":
            data_list = [['ward Name' , 'Health Post Name' , "CHV-ASHA" , "CHV-ASHA Mobile number" ,'Families Enrolled'  , 'Citizens Enrolled' , 'CBAC Filled' ,
                        'Citizens 60 years + enrolled', 'Citizens 30 years + enrolled' , "Males Enrolled" ,  "Females Enrolled" ,  "Transgender Enrolled",
                        "ABHA ID Generated" , 'Diabetes' , 'Hypertension' ,  'Oral Cancer' , 'Cervical cancer' , 'COPD' , 'Eye Disorder' ,
                        'ENT Disorder' ,  'Asthma' , 'Alzheimers' ,  'TB' , 'leprosy','Breast cancer' , 'Other Communicable' ,
                        'Blood collected at home' , 'blood collected at center' , 'Blood Collection Denied By AMO' ,  'Blood Collection Denied By Citizen',
                        'Total Reports Generated' , 'Tests Assigned' ,
                        'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment' ,
                        'Referral to HBT polyclinic for Physician consultation',
                        'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                        'Referral to Medical College for management of Complication',
                        'Referral to Private facility',
                        'Vulnerable Citizen' ]]

        header1 = {'Citizen Details':len(data_list[0][0:13]), 'Dieases Suspected' : len(data_list[0][13:25]),
                   'Blood Collection' : len(data_list[0][25:31]), 'Referrals' : len(data_list[0][31:])}

        if healthpost_id:

            try:
                healthpost = healthPost.objects.get(pk=healthpost_id)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists with ID %d"%(id),
                    "status":"error"
                } , status= 400 )

            healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            healthpost_name = healthpost.healthPostName

            if not healthpost_related_user:
                return Response({
                    "message":"No data found for healthpost %s"%(healthpost_name),
                    "status":"error"} , status = 400)

            health_workers = CustomUser.objects.filter(groups__name = group_name , userSections__healthPost__id = healthpost_id).distinct().order_by("name")

            for healthworker in health_workers:

                # Distinct and common queries of survey data
                healthpost_queryset = healthpost_related_user.filter(familySurveyor__id=healthworker.id)
                familySurvey_queryset = self.FamilySurvey_count.filter(user__id=healthworker.id)

                # Healthppost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append([healthpost.ward.wardName, healthpost.healthPostName, healthworker.name, healthworker.phoneNumber,
                                combined_survey_data["total_family_count"], combined_survey_data["total_citizen_count"],
                                combined_survey_data["total_cbac_count"], combined_survey_data["citizen_above_60"],
                                combined_survey_data["citizen_above_30"], combined_survey_data["male"],
                                combined_survey_data["female"], combined_survey_data["transgender"], combined_survey_data["total_AbhaCreated"],
                                suspected_disease_counts["diabetes"], combined_survey_data["hypertension"],
                                suspected_disease_counts["oral_Cancer"], suspected_disease_counts["cervical_cancer"],
                                suspected_disease_counts["copd"], suspected_disease_counts["eye_disorder"],
                                suspected_disease_counts["ent_disorder"], suspected_disease_counts["asthama"],
                                suspected_disease_counts["Alzheimers"], suspected_disease_counts["tb"],
                                suspected_disease_counts["leprosy"], suspected_disease_counts["breast_cancer"],
                                suspected_disease_counts["other_communicable_dieases"],
                                combined_survey_data["blood_collected_home"], combined_survey_data["blood_collected_center"],
                                combined_survey_data["denied_by_mo_count"] ,  combined_survey_data["denied_by_mo_individual"],
                                combined_survey_data["TestReportGenerated"], combined_survey_data["total_LabTestAdded"],
                                combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                combined_survey_data["total_vulnerable"]])

                wb = openpyxl.Workbook()
                ws = wb.active
                self.add_headers(ws, header1 )
                for row in data_list:
                    ws.append(row)

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(healthpost_name+"_data_"+today)
            wb.save(response)
            return response

        elif wardId:

            try:
                ward_obj = ward.objects.get(pk=wardId)
            except ward.DoesNotExist:
                return Response({
                    "message":"No ward exists with ID %d"%(wardId),
                    "status":"error"
                } , status = 400)

            ward_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=wardId).exists()

            if not ward_related_user:
                return Response({
                    "message":"No data found for ward %s"%(ward_obj.wardName),
                    "status":"error"
                } , status = 400)

            healthposts = healthPost.objects.filter(ward__id = wardId)

            for healthpost in healthposts:

                health_workers = CustomUser.objects.filter(groups__name = group_name , userSections__healthPost__id = healthpost.id).distinct().order_by("name")

                for healthworker in health_workers:

                    # Common querysets of survey
                    ward_queryset = self.get_queryset().filter(familySurveyor__id=healthworker.id)
                    familySurvey_queryset = self.FamilySurvey_count.filter(user__id=healthworker.id)

                    # Ward related survey data
                    ward_data = ward_queryset.aggregate(
                        total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                        total_citizen_count=Count('id', distinct=True),
                        total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                        male=Count('id', filter=Q(gender="M"), distinct=True),
                        female=Count('id', filter=Q(gender="F"), distinct=True),
                        transgender=Count('id', filter=Q(gender="O"), distinct=True),
                        citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                        citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                        total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                        blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                        blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                        denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                        denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                        Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                        Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                        Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                        Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                        Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                        hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                        total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                        TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                    )

                    # Aggregate counts for familySurvey_queryset
                    familySurvey_data = familySurvey_queryset.aggregate(
                        total_family_count=Count('id', distinct=True),
                    )

                    combined_survey_data = {**ward_data, **familySurvey_data}

                    Questionnaire_queryset = ward_queryset.filter(Questionnaire__isnull=False)
                    suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                    data_list.append([healthpost.ward.wardName, healthpost.healthPostName, healthworker.name, healthworker.phoneNumber,
                                combined_survey_data["total_family_count"], combined_survey_data["total_citizen_count"],
                                combined_survey_data["total_cbac_count"], combined_survey_data["citizen_above_60"],
                                combined_survey_data["citizen_above_30"], combined_survey_data["male"],
                                combined_survey_data["female"], combined_survey_data["transgender"], combined_survey_data["total_AbhaCreated"],
                                suspected_disease_counts["diabetes"], combined_survey_data["hypertension"],
                                suspected_disease_counts["oral_Cancer"], suspected_disease_counts["cervical_cancer"],
                                suspected_disease_counts["copd"], suspected_disease_counts["eye_disorder"],
                                suspected_disease_counts["ent_disorder"], suspected_disease_counts["asthama"],
                                suspected_disease_counts["Alzheimers"], suspected_disease_counts["tb"],
                                suspected_disease_counts["leprosy"], suspected_disease_counts["breast_cancer"],
                                suspected_disease_counts["other_communicable_dieases"],
                                combined_survey_data["blood_collected_home"], combined_survey_data["blood_collected_center"],
                                combined_survey_data["denied_by_mo_count"] ,  combined_survey_data["denied_by_mo_individual"],
                                combined_survey_data["TestReportGenerated"], combined_survey_data["total_LabTestAdded"],
                                combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                combined_survey_data["total_vulnerable"]])


                    wb = openpyxl.Workbook()
                    ws = wb.active
                    self.add_headers(ws, header1 )
                    for row in data_list:
                        ws.append(row)


            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(ward_obj.wardName+"_data_"+today)
            wb.save(response)
            return response

        else:
            return Response({'status': 'error',
                              'message': "Please select Ward or HealthPost"}, status = 400)


class AdminDashboardView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated , IsAdmin|IsViewAdmin)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def get(self, request ,  *args, **kwargs):

        healthpost_id = request.query_params.get('healthpost_id', None)
        wardId = request.query_params.get('wardId', None)

        CHV_ASHA_count = self.CustomUser_queryset.filter(groups__name='CHV-ASHA').count()
        MO_count = self.CustomUser_queryset.filter(groups__name='mo').count()
        ANM_count = self.CustomUser_queryset.filter(groups__name='healthworker').count()
        today = timezone.now().date()

        if healthpost_id:

            # Distinct and common queries of survey data
            healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            healthpost_data = healthpost_queryset.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                todays_citizen_count=Count('id', filter=Q(created_date__date=today), distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                vulnerable_70_Years=Count('id', filter=Q(vulnerable_choices__choice='70+ Years'), distinct=True),
                vulnerable_Physically_handicapped=Count('id', filter=Q(vulnerable_choices__choice='Physically Handicapped'), distinct=True),
                vulnerable_completely_paralyzed_or_on_bed=Count('id', filter=Q(vulnerable_choices__choice='Completely Paralyzed or On bed'), distinct=True),
                vulnerable_elderly_and_alone_at_home=Count('id', filter=Q(vulnerable_choices__choice='Elderly and alone at home'), distinct=True),
                vulnerable_any_other_reason=Count('id', filter=Q(vulnerable_choices__choice='Any other reason'), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                partial_survey_count=Count('id', filter=Q(partialSubmit=True), distinct=True),
                total_family_count=Count('id', distinct=True),
                today_family_count=Count('id', filter=Q(created_date__date=today), distinct=True)
            )

            combined_survey_data = {**healthpost_data, **familySurvey_data}

            Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

        elif wardId :

            # Distinct and common queries of survey data
            ward_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=wardId).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward__id=wardId)

            # Ward related survey data
            ward_data = ward_queryset.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                todays_citizen_count=Count('id', filter=Q(created_date__date=today), distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                vulnerable_70_Years=Count('id', filter=Q(vulnerable_choices__choice='70+ Years'), distinct=True),
                vulnerable_Physically_handicapped=Count('id', filter=Q(vulnerable_choices__choice='Physically Handicapped'), distinct=True),
                vulnerable_completely_paralyzed_or_on_bed=Count('id', filter=Q(vulnerable_choices__choice='Completely Paralyzed or On bed'), distinct=True),
                vulnerable_elderly_and_alone_at_home=Count('id', filter=Q(vulnerable_choices__choice='Elderly and alone at home'), distinct=True),
                vulnerable_any_other_reason=Count('id', filter=Q(vulnerable_choices__choice='Any other reason'), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                partial_survey_count=Count('id', filter=Q(partialSubmit=True), distinct=True),
                total_family_count=Count('id', distinct=True),
                today_family_count=Count('id', filter=Q(created_date__date=today), distinct=True)
            )

            combined_survey_data = {**ward_data, **familySurvey_data}

            Questionnaire_queryset = ward_queryset.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

        else:
            # survey data
            survey_data = self.get_queryset().aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                todays_citizen_count=Count('id', filter=Q(created_date__date=today), distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                vulnerable_70_Years=Count('id', filter=Q(vulnerable_choices__choice='70+ Years'), distinct=True),
                vulnerable_Physically_handicapped=Count('id', filter=Q(vulnerable_choices__choice='Physically Handicapped'), distinct=True),
                vulnerable_completely_paralyzed_or_on_bed=Count('id', filter=Q(vulnerable_choices__choice='Completely Paralyzed or On bed'), distinct=True),
                vulnerable_elderly_and_alone_at_home=Count('id', filter=Q(vulnerable_choices__choice='Elderly and alone at home'), distinct=True),
                vulnerable_any_other_reason=Count('id', filter=Q(vulnerable_choices__choice='Any other reason'), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = self.FamilySurvey_count.aggregate(
                partial_survey_count=Count('id', filter=Q(partialSubmit=True), distinct=True),
                total_family_count=Count('id', distinct=True),
                today_family_count=Count('id', filter=Q(created_date__date=today), distinct=True)
            )

            combined_survey_data = {**survey_data, **familySurvey_data}

            Questionnaire_queryset = self.get_queryset().filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

        return Response({
            'CHV_ASHA_count' : CHV_ASHA_count,
            'MO_count' : MO_count,
            'ANM_count' : ANM_count,
            'total_count' : combined_survey_data["total_citizen_count"],
            'todays_count' : combined_survey_data["todays_citizen_count"],
            'partial_survey_count' : combined_survey_data["partial_survey_count"],
            'total_family_count' : combined_survey_data["total_family_count"],
            'today_family_count' : combined_survey_data["today_family_count"],
            'total_cbac_count' : combined_survey_data["total_cbac_count"],
            'citizen_above_60' : combined_survey_data["citizen_above_60"],
            'citizen_above_30' : combined_survey_data["citizen_above_30"],
            'total_AbhaCreated' : combined_survey_data["total_AbhaCreated"],
            'TestReportGenerated' : combined_survey_data["TestReportGenerated"],
            'total_LabTestAdded' : combined_survey_data["total_LabTestAdded"],
            "male" : combined_survey_data["male"],
            "female" : combined_survey_data["female"],
            "transgender" : combined_survey_data["transgender"],
            'hypertension' : combined_survey_data["hypertension"],
            **suspected_disease_counts,
            'blood_collected_home' : combined_survey_data["blood_collected_home"],
            'blood_collected_center' : combined_survey_data["blood_collected_center"],
            'denieded_by_mo_count' : combined_survey_data["denied_by_mo_count"],
            'denieded_by_mo_individual' : combined_survey_data["denied_by_mo_individual"],
            'Referral_choice_Referral_to_Mun_Dispensary' : combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
            'Referral_choice_Referral_to_HBT_polyclinic': combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
            'Referral_choice_Referral_to_Peripheral_Hospital': combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
            'Referral_choice_Referral_to_Medical_College': combined_survey_data["Referral_choice_Referral_to_Medical_College"],
            'Referral_choice_Referral_to_Private_facility': combined_survey_data["Referral_choice_Referral_to_Private_facility"],
            'total_vulnerabel' : combined_survey_data["total_vulnerable"],
            'vulnerabel_70_Years' : combined_survey_data["vulnerable_70_Years"],
            'vulnerabel_Physically_handicapped' : combined_survey_data["vulnerable_Physically_handicapped"],
            'vulnerabel_completely_paralyzed_or_on_bed' : combined_survey_data["vulnerable_completely_paralyzed_or_on_bed"],
            'vulnerabel_elderly_and_alone_at_home' : combined_survey_data["vulnerable_elderly_and_alone_at_home"],
            'vulnerabel_any_other_reason' : combined_survey_data["vulnerable_any_other_reason"]}, status= 200)

class AdminDashboardExportView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated , IsAdmin|IsViewAdmin)
    FamilySurvey_count = familyHeadDetails.objects.all().order_by('user__userSections__healthPost__ward')
    queryset = familyMembers.objects.all()
    healthposts = healthPost.objects.all().order_by('ward__wardName')


    def add_headers(self, sheet, *args):
        """This function takes the sheet and add merged headers to that sheet"""
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def get(self, request):
        healthpost_id = request.query_params.get('healthpost_id', None)
        wardId = request.query_params.get('wardId', None)
        today = datetime.today().strftime('%d-%m-%Y')

        data_list = [['ward Name' , 'Health Post Name' , 'Families Enrolled'  , 'Citizens Enrolled' , 'CBAC Filled' ,
                    'Citizens 60 years + enrolled', 'Citizens 30 years + enrolled' , "Males Enrolled" ,  "Females Enrolled" ,  "Transgender Enrolled",
                    "ABHA ID Generated" , 'Diabetes' , 'Hypertension' ,  'Oral Cancer' , 'Cervical cancer' , 'COPD' , 'Eye Disorder' ,
                    'ENT Disorder' ,  'Asthma' , 'Alzheimers' ,  'TB' , 'Leprosy', 'Breast cancer' , 'Other Communicable' ,
                    'Blood collected at home' , 'blood collected at center' , 'Blood Collection Denied By AMO' ,  'Blood Collection Denied By Citizen',
                    'Total Reports Generated' , 'Tests Assigned' ,
                    'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment' ,
                    'Referral to HBT polyclinic for Physician consultation',
                    'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                    'Referral to Medical College for management of Complication',
                    'Referral to Private facility',
                    'Vulnerable Citizen' ]]

        header1 = {'Citizen Details':len(data_list[0][0:11]) ,'Dieases Suspected': len(data_list[0][12:24]),
                   'Blood Collection' : len(data_list[0][24:30]) , 'Referrals' : len(data_list[0][30:]) }

        if healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists with ID %d"%(id),
                    "status":"error"
                } , status = 400)

            healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            healthpost_name = healthpost.healthPostName

            if not healthpost_related_user:
                return Response({
                    "message":"No data found for healthpost %s"%(healthpost_name),
                    "status":"error"
                } , status = 400)

            # Distinct and common queries of survey data
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            healthpost_data = healthpost_related_user.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                total_family_count=Count('id', distinct=True),
            )

            combined_survey_data = {**healthpost_data, **familySurvey_data}

            Questionnaire_queryset = healthpost_related_user.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

            data_list.append([healthpost.ward.wardName, healthpost.healthPostName,
                            combined_survey_data["total_family_count"], combined_survey_data["total_citizen_count"],
                            combined_survey_data["total_cbac_count"], combined_survey_data["citizen_above_60"],
                            combined_survey_data["citizen_above_30"], combined_survey_data["male"],
                            combined_survey_data["female"], combined_survey_data["transgender"], combined_survey_data["total_AbhaCreated"],
                            suspected_disease_counts["diabetes"], combined_survey_data["hypertension"],
                            suspected_disease_counts["oral_Cancer"], suspected_disease_counts["cervical_cancer"],
                            suspected_disease_counts["copd"], suspected_disease_counts["eye_disorder"],
                            suspected_disease_counts["ent_disorder"], suspected_disease_counts["asthama"],
                            suspected_disease_counts["Alzheimers"], suspected_disease_counts["tb"],
                            suspected_disease_counts["leprosy"], suspected_disease_counts["breast_cancer"],
                            suspected_disease_counts["other_communicable_dieases"],
                            combined_survey_data["blood_collected_home"], combined_survey_data["blood_collected_center"],
                            combined_survey_data["denied_by_mo_count"] ,  combined_survey_data["denied_by_mo_individual"],
                            combined_survey_data["TestReportGenerated"], combined_survey_data["total_LabTestAdded"],
                            combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                            combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                            combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                            combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                            combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                            combined_survey_data["total_vulnerable"]])

            wb = openpyxl.Workbook()
            ws = wb.active
            self.add_headers(ws, header1 )
            for row in data_list:
                # print(row)
                ws.append(row)

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(healthpost_name+"_data_"+today)
            wb.save(response)
            return response

        elif wardId:
            try:
                ward_name = ward.objects.get(pk=wardId)
            except ward.DoesNotExist:
                return Response({
                    "message":"No ward exists with ID %s"%(wardId),
                    "status":"error"
                } , status = 400)

            ward_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=wardId).exists()

            if not ward_related_user:
                return Response({
                    "message":"No data found for ward %s"%(ward_name.wardName),
                    "status":"error"
                } , status = 400)

            healthposts = self.healthposts.filter(ward__id=wardId)

            for healthpost in healthposts:
                # Common querysets of survey
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthpost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append([healthpost.ward.wardName, healthpost.healthPostName,
                                combined_survey_data["total_family_count"], combined_survey_data["total_citizen_count"],
                                combined_survey_data["total_cbac_count"], combined_survey_data["citizen_above_60"],
                                combined_survey_data["citizen_above_30"], combined_survey_data["male"],
                                combined_survey_data["female"], combined_survey_data["transgender"], combined_survey_data["total_AbhaCreated"],
                                suspected_disease_counts["diabetes"], combined_survey_data["hypertension"],
                                suspected_disease_counts["oral_Cancer"], suspected_disease_counts["cervical_cancer"],
                                suspected_disease_counts["copd"], suspected_disease_counts["eye_disorder"],
                                suspected_disease_counts["ent_disorder"], suspected_disease_counts["asthama"],
                                suspected_disease_counts["Alzheimers"], suspected_disease_counts["tb"],
                                suspected_disease_counts["leprosy"], suspected_disease_counts["breast_cancer"],
                                suspected_disease_counts["other_communicable_dieases"],
                                combined_survey_data["blood_collected_home"], combined_survey_data["blood_collected_center"],
                                combined_survey_data["denied_by_mo_count"] ,  combined_survey_data["denied_by_mo_individual"],
                                combined_survey_data["TestReportGenerated"], combined_survey_data["total_LabTestAdded"],
                                combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                combined_survey_data["total_vulnerable"]])

            wb = openpyxl.Workbook()
            ws = wb.active
            self.add_headers(ws, header1 )
            for row in data_list:
                ws.append(row)

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("Ward_"+ward_name.wardName+"_data_"+today)
            wb.save(response)
            return response
        else:
            for healthpost in self.healthposts:
                # Distinct and common queries of survey data
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthppost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append([healthpost.ward.wardName, healthpost.healthPostName,
                                combined_survey_data["total_family_count"], combined_survey_data["total_citizen_count"],
                                combined_survey_data["total_cbac_count"], combined_survey_data["citizen_above_60"],
                                combined_survey_data["citizen_above_30"], combined_survey_data["male"],
                                combined_survey_data["female"], combined_survey_data["transgender"], combined_survey_data["total_AbhaCreated"],
                                suspected_disease_counts["diabetes"], combined_survey_data["hypertension"],
                                suspected_disease_counts["oral_Cancer"], suspected_disease_counts["cervical_cancer"],
                                suspected_disease_counts["copd"], suspected_disease_counts["eye_disorder"],
                                suspected_disease_counts["ent_disorder"], suspected_disease_counts["asthama"],
                                suspected_disease_counts["Alzheimers"], suspected_disease_counts["tb"],
                                suspected_disease_counts["leprosy"], suspected_disease_counts["breast_cancer"],
                                suspected_disease_counts["other_communicable_dieases"],
                                combined_survey_data["blood_collected_home"], combined_survey_data["blood_collected_center"],
                                combined_survey_data["denied_by_mo_count"] ,  combined_survey_data["denied_by_mo_individual"],
                                combined_survey_data["TestReportGenerated"], combined_survey_data["total_LabTestAdded"],
                                combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                combined_survey_data["total_vulnerable"]])

            wb = openpyxl.Workbook()
            ws = wb.active
            self.add_headers(ws, header1)
            for row in data_list:
                ws.append(row)

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("All_Ward_data_"+today)
            wb.save(response)
            return response

class MOHDashboardTabView(generics.GenericAPIView): #Modified

    permission_classes = (IsAuthenticated, IsMOH)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    pagination_class = CustomPageNumberPagination  # Add pagination class

    def get(self, request):

        healthpost_id = request.query_params.get('healthpost_id', None)
        wardId = request.user.ward_id

        data_list =[]

        if healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id, ward_id=wardId)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists",
                    "status":"error"
                } , status= 400 )

            healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()

            if not healthpost_related_user:
                data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                "total_family_count":0 , "total_citizen_count":0 ,
                "total_cbac_count":0 , "citizen_above_60":0 ,
                "citizen_above_30":0 , "male":0 , "female":0 ,"transgender":0 ,"total_AbhaCreated":0 ,
                "total_diabetes":0 , "hypertension":0 , "total_oral_cancer":0 , "total_cervical_cancer":0 , "total_COPD_count":0 , "total_eye_problem":0 ,
                "total_ent_disorder":0 , "total_asthma":0, "total_Alzheimers":0, "total_tb_count":0 ,"total_leprosy":0 , "total_breast_cancer":0 , "total_communicable":0 ,
                "blood_collected_home":0 , "blood_collected_center":0 ,  "denieded_by_mo_count":0  ,  "denieded_by_mo_individual":0 , "TestReportGenerated":0 , "total_LabTestAdded":0,
                "Referral_choice_Referral_to_Mun_Dispensary":0 ,
                "Referral_choice_Referral_to_HBT_polyclinic":0,
                "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0 ,
                "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerabel":0})

                return Response({
                    "message":"Successfully Fetched",
                    "status":"success","results":data_list} , status = 200)

            # Common querysets of survey
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthpost related survey data
            healthpost_data = healthpost_related_user.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                total_family_count=Count('id', distinct=True),
            )

            combined_survey_data = {**healthpost_data, **familySurvey_data}

            Questionnaire_queryset = healthpost_related_user.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

            data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                            "total_family_count":combined_survey_data["total_family_count"], "total_citizen_count":combined_survey_data["total_citizen_count"],
                            "total_cbac_count":combined_survey_data["total_cbac_count"], "citizen_above_60":combined_survey_data["citizen_above_60"],
                            "citizen_above_30":combined_survey_data["citizen_above_30"], "male":combined_survey_data["male"],
                            "female":combined_survey_data["female"], "transgender":combined_survey_data["transgender"], "total_AbhaCreated":combined_survey_data["total_AbhaCreated"],
                            "total_diabetes":suspected_disease_counts["diabetes"], "hypertension":combined_survey_data["hypertension"], "total_oral_cancer":suspected_disease_counts["oral_Cancer"],
                            "total_cervical_cancer":suspected_disease_counts["cervical_cancer"], "total_COPD_count":suspected_disease_counts["copd"],
                            "total_eye_problem":suspected_disease_counts["eye_disorder"], "total_ent_disorder":suspected_disease_counts["ent_disorder"],
                            "total_asthma":suspected_disease_counts["asthama"], "total_Alzheimers":suspected_disease_counts["Alzheimers"],
                            "total_tb_count":suspected_disease_counts["tb"], "total_leprosy":suspected_disease_counts["leprosy"],
                            "total_breast_cancer":suspected_disease_counts["breast_cancer"], "total_communicable":suspected_disease_counts["other_communicable_dieases"],
                            "blood_collected_home":combined_survey_data["blood_collected_home"], "blood_collected_center":combined_survey_data["blood_collected_center"],
                            "denieded_by_mo_count":combined_survey_data["denied_by_mo_count"] , "denieded_by_mo_individual":combined_survey_data["denied_by_mo_individual"],
                            "TestReportGenerated":combined_survey_data["TestReportGenerated"], "total_LabTestAdded":combined_survey_data["total_LabTestAdded"],
                            "Referral_choice_Referral_to_Mun_Dispensary":combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                            "Referral_choice_Referral_to_HBT_polyclinic":combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                            "Referral_choice_Referral_to_Peripheral_Hospital":combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                            "Referral_choice_Referral_to_Medical_College":combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                            "Referral_choice_Referral_to_Private_facility":combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                            "total_vulnerabel":combined_survey_data["total_vulnerable"]})

            page = self.paginate_queryset(data_list)
            if page is not None:
                return self.get_paginated_response(page)
            return Response({
                    "message":"Successfully Fetched",
                    "status":"success",
                    "data":data_list
                }, status = 200)
        else:
            healthposts = healthPost.objects.filter(ward__id = wardId)

            for healthpost in healthposts:

                # Common querysets of survey
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthpost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                                "total_family_count":combined_survey_data["total_family_count"], "total_citizen_count":combined_survey_data["total_citizen_count"],
                                "total_cbac_count":combined_survey_data["total_cbac_count"], "citizen_above_60":combined_survey_data["citizen_above_60"],
                                "citizen_above_30":combined_survey_data["citizen_above_30"], "male":combined_survey_data["male"],
                                "female":combined_survey_data["female"], "transgender":combined_survey_data["transgender"], "total_AbhaCreated":combined_survey_data["total_AbhaCreated"],
                                "total_diabetes":suspected_disease_counts["diabetes"], "hypertension":combined_survey_data["hypertension"], "total_oral_cancer":suspected_disease_counts["oral_Cancer"],
                                "total_cervical_cancer":suspected_disease_counts["cervical_cancer"], "total_COPD_count":suspected_disease_counts["copd"],
                                "total_eye_problem":suspected_disease_counts["eye_disorder"], "total_ent_disorder":suspected_disease_counts["ent_disorder"],
                                "total_asthma":suspected_disease_counts["asthama"], "total_Alzheimers":suspected_disease_counts["Alzheimers"],
                                "total_tb_count":suspected_disease_counts["tb"], "total_leprosy":suspected_disease_counts["leprosy"],
                                "total_breast_cancer":suspected_disease_counts["breast_cancer"], "total_communicable":suspected_disease_counts["other_communicable_dieases"],
                                "blood_collected_home":combined_survey_data["blood_collected_home"], "blood_collected_center":combined_survey_data["blood_collected_center"],
                                "denieded_by_mo_count":combined_survey_data["denied_by_mo_count"] , "denieded_by_mo_individual":combined_survey_data["denied_by_mo_individual"],
                                "TestReportGenerated":combined_survey_data["TestReportGenerated"], "total_LabTestAdded":combined_survey_data["total_LabTestAdded"],
                                "Referral_choice_Referral_to_Mun_Dispensary":combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                "Referral_choice_Referral_to_HBT_polyclinic":combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                "Referral_choice_Referral_to_Peripheral_Hospital":combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                "Referral_choice_Referral_to_Medical_College":combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                "Referral_choice_Referral_to_Private_facility":combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                "total_vulnerabel":combined_survey_data["total_vulnerable"]})

            page = self.paginate_queryset(data_list)
            if page is not None:
                return self.get_paginated_response(page)
            return Response({
                    "message":"Successfully Fetched",
                    "status":"success",
                    "data":data_list
                } , status = 200)

class AdminDashboardTabView(generics.GenericAPIView): #Modified
    permission_classes = (IsAuthenticated , IsAdmin|IsViewAdmin)
    FamilySurvey_count = familyHeadDetails.objects.all().order_by('user__userSections__healthPost__ward')
    queryset = familyMembers.objects.all()
    healthposts = healthPost.objects.all().order_by('ward__wardName')
    pagination_class = CustomPageNumberPagination  # Add pagination class

    def get(self, request):

        healthpost_id = request.query_params.get('healthpost_id', None)
        wardId = request.query_params.get('wardId', None)

        data_list = [] # List for storing survey data
        today = datetime.today().strftime('%d-%m-%Y')

        if wardId and not healthpost_id:

            try:
                ward_name = ward.objects.get(pk=wardId)
            except ward.DoesNotExist:
                return Response({
                    "message":"No ward exists with ID %s"%(wardId),
                    "status":"error"
                } , status = 400)

            ward_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost__ward__id=wardId).exists()

            if not ward_related_user:
                data_list.append({"wardName":ward_name.wardName, "healthPostName":"" ,
                "total_family_count":0 ,
                "total_citizen_count":0 ,
                "total_cbac_count":0 ,
                "citizen_above_60":0 ,
                "citizen_above_30":0 ,
                "male":0 ,
                "female":0 ,
                "transgender":0 ,
                "total_AbhaCreated":0 ,
                "total_diabetes":0 , "hypertension":0 , "total_oral_cancer":0 , "total_cervical_cancer":0 , "total_COPD_count":0 , "total_eye_problem":0 ,
                "total_ent_disorder":0 , "total_asthma":0, "total_Alzheimers":0, "total_tb_count":0 ,"total_leprosy":0 , "total_breast_cancer":0 , "total_communicable":0 ,
                "blood_collected_home":0 , "blood_collected_center":0 ,  "denieded_by_mo_count":0  ,  "denieded_by_mo_individual":0 , "TestReportGenerated":0 , "total_LabTestAdded":0,
                "Referral_choice_Referral_to_Mun_Dispensary":0 ,
                "Referral_choice_Referral_to_HBT_polyclinic":0,
                "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0 ,
                "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerabel":0})

                return Response({
                    "message":"Successfully Fetched",
                    "status":"success","results":data_list} , status = 200)

            healthposts = self.healthposts.filter(ward__id=wardId)

            for healthpost in healthposts:
                # Common querysets of survey
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthpost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                                "total_family_count":combined_survey_data["total_family_count"], "total_citizen_count":combined_survey_data["total_citizen_count"],
                                "total_cbac_count":combined_survey_data["total_cbac_count"], "citizen_above_60":combined_survey_data["citizen_above_60"],
                                "citizen_above_30":combined_survey_data["citizen_above_30"], "male":combined_survey_data["male"],
                                "female":combined_survey_data["female"], "transgender":combined_survey_data["transgender"], "total_AbhaCreated":combined_survey_data["total_AbhaCreated"],
                                "total_diabetes":suspected_disease_counts["diabetes"], "hypertension":combined_survey_data["hypertension"], "total_oral_cancer":suspected_disease_counts["oral_Cancer"],
                                "total_cervical_cancer":suspected_disease_counts["cervical_cancer"], "total_COPD_count":suspected_disease_counts["copd"],
                                "total_eye_problem":suspected_disease_counts["eye_disorder"], "total_ent_disorder":suspected_disease_counts["ent_disorder"],
                                "total_asthma":suspected_disease_counts["asthama"], "total_Alzheimers":suspected_disease_counts["Alzheimers"],
                                "total_tb_count":suspected_disease_counts["tb"], "total_leprosy":suspected_disease_counts["leprosy"],
                                "total_breast_cancer":suspected_disease_counts["breast_cancer"], "total_communicable":suspected_disease_counts["other_communicable_dieases"],
                                "blood_collected_home":combined_survey_data["blood_collected_home"], "blood_collected_center":combined_survey_data["blood_collected_center"],
                                "denieded_by_mo_count":combined_survey_data["denied_by_mo_count"] , "denieded_by_mo_individual":combined_survey_data["denied_by_mo_individual"],
                                "TestReportGenerated":combined_survey_data["TestReportGenerated"], "total_LabTestAdded":combined_survey_data["total_LabTestAdded"],
                                "Referral_choice_Referral_to_Mun_Dispensary":combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                "Referral_choice_Referral_to_HBT_polyclinic":combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                "Referral_choice_Referral_to_Peripheral_Hospital":combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                "Referral_choice_Referral_to_Medical_College":combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                "Referral_choice_Referral_to_Private_facility":combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                "total_vulnerabel":combined_survey_data["total_vulnerable"]})

            page = self.paginate_queryset(data_list)
            if page is not None:
                return self.get_paginated_response(page)

            return Response({
                "status":"success",
                "message":"Successfully Fetched",
                "data":data_list}, status=200)

        elif wardId and healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id, ward_id=wardId)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists with ID %s"%(healthpost_id),
                    "status":"error",
                } , status = 400)

            healthpost_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost_id=healthpost_id).distinct()
            healthpost_name = healthpost.healthPostName

            if not healthpost_related_user:
                data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost_name ,
                "total_family_count":0 ,
                "total_citizen_count":0 ,
                "total_cbac_count":0 ,
                "citizen_above_60":0 ,
                "citizen_above_30":0 ,
                "male":0 ,
                "female":0 ,
                "transgender":0 ,
                "total_AbhaCreated":0 ,
                "total_diabetes":0 , "hypertension":0 , "total_oral_cancer":0 , "total_cervical_cancer":0 , "total_COPD_count":0 , "total_eye_problem":0 ,
                "total_ent_disorder":0 , "total_asthma":0, "total_Alzheimers":0, "total_tb_count":0 ,"total_leprosy":0 , "total_breast_cancer":0 , "total_communicable":0 ,
                "blood_collected_home":0 , "blood_collected_center":0 ,  "denieded_by_mo_count":0  ,  "denieded_by_mo_individual":0 , "TestReportGenerated":0 , "total_LabTestAdded":0,
                "Referral_choice_Referral_to_Mun_Dispensary":0 ,
                "Referral_choice_Referral_to_HBT_polyclinic":0,
                "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0 ,
                "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerabel":0})

                return Response({
                    "message":"Successfully Fetched",
                    "status":"success","results":data_list}, status = 200)

            # Common querysets of survey
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost_id)

            # Healthpost related survey data
            healthpost_data = healthpost_related_user.aggregate(
                total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                total_citizen_count=Count('id', distinct=True),
                total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                male=Count('id', filter=Q(gender="M"), distinct=True),
                female=Count('id', filter=Q(gender="F"), distinct=True),
                transgender=Count('id', filter=Q(gender="O"), distinct=True),
                citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
            )

            # Aggregate counts for familySurvey_queryset
            familySurvey_data = familySurvey_queryset.aggregate(
                total_family_count=Count('id', distinct=True),
            )

            combined_survey_data = {**healthpost_data, **familySurvey_data}

            Questionnaire_queryset = healthpost_related_user.filter(Questionnaire__isnull=False)
            suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

            data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                            "total_family_count":combined_survey_data["total_family_count"], "total_citizen_count":combined_survey_data["total_citizen_count"],
                            "total_cbac_count":combined_survey_data["total_cbac_count"], "citizen_above_60":combined_survey_data["citizen_above_60"],
                            "citizen_above_30":combined_survey_data["citizen_above_30"], "male":combined_survey_data["male"],
                            "female":combined_survey_data["female"], "transgender":combined_survey_data["transgender"], "total_AbhaCreated":combined_survey_data["total_AbhaCreated"],
                            "total_diabetes":suspected_disease_counts["diabetes"], "hypertension":combined_survey_data["hypertension"], "total_oral_cancer":suspected_disease_counts["oral_Cancer"],
                            "total_cervical_cancer":suspected_disease_counts["cervical_cancer"], "total_COPD_count":suspected_disease_counts["copd"],
                            "total_eye_problem":suspected_disease_counts["eye_disorder"], "total_ent_disorder":suspected_disease_counts["ent_disorder"],
                            "total_asthma":suspected_disease_counts["asthama"], "total_Alzheimers":suspected_disease_counts["Alzheimers"],
                            "total_tb_count":suspected_disease_counts["tb"], "total_leprosy":suspected_disease_counts["leprosy"],
                            "total_breast_cancer":suspected_disease_counts["breast_cancer"], "total_communicable":suspected_disease_counts["other_communicable_dieases"],
                            "blood_collected_home":combined_survey_data["blood_collected_home"], "blood_collected_center":combined_survey_data["blood_collected_center"],
                            "denieded_by_mo_count":combined_survey_data["denied_by_mo_count"] , "denieded_by_mo_individual":combined_survey_data["denied_by_mo_individual"],
                            "TestReportGenerated":combined_survey_data["TestReportGenerated"], "total_LabTestAdded":combined_survey_data["total_LabTestAdded"],
                            "Referral_choice_Referral_to_Mun_Dispensary":combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                            "Referral_choice_Referral_to_HBT_polyclinic":combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                            "Referral_choice_Referral_to_Peripheral_Hospital":combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                            "Referral_choice_Referral_to_Medical_College":combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                            "Referral_choice_Referral_to_Private_facility":combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                            "total_vulnerabel":combined_survey_data["total_vulnerable"]})

            page = self.paginate_queryset(data_list)
            if page is not None:
                return self.get_paginated_response(page)


            return Response({
                    "status":"success",
                    "message":"Successfully Fetched",
                    "data":data_list} , status = 200)

        else:
            for healthpost in self.healthposts:

                # Distinct and common queries of survey data
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost_id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost.id)

                # Healthppost related survey data
                healthpost_data = healthpost_queryset.aggregate(
                    total_AbhaCreated=Count('id', filter=Q(isAbhaCreated=True), distinct=True),
                    total_citizen_count=Count('id', distinct=True),
                    total_cbac_count=Count('id', filter=Q(age__gte=30, cbacRequired=True), distinct=True),
                    male=Count('id', filter=Q(gender="M"), distinct=True),
                    female=Count('id', filter=Q(gender="F"), distinct=True),
                    transgender=Count('id', filter=Q(gender="O"), distinct=True),
                    citizen_above_60=Count('id', filter=Q(age__gte=60), distinct=True),
                    citizen_above_30=Count('id', filter=Q(age__gte=30), distinct=True),
                    total_vulnerable=Count('id', filter=Q(vulnerable=True), distinct=True),
                    blood_collected_home=Count('id', filter=Q(bloodCollectionLocation='Home'), distinct=True),
                    blood_collected_center=Count('id', filter=Q(bloodCollectionLocation='Center'), distinct=True),
                    denied_by_mo_count=Count('id', filter=Q(bloodCollectionLocation='AMO'), distinct=True),
                    denied_by_mo_individual=Count('id', filter=Q(bloodCollectionLocation='Individual Itself'), distinct=True),
                    Referral_choice_Referral_to_Mun_Dispensary=Count('id', filter=Q(referels__choice='Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment'), distinct=True),
                    Referral_choice_Referral_to_HBT_polyclinic=Count('id', filter=Q(referels__choice='Referral to HBT polyclinic for physician consultation'), distinct=True),
                    Referral_choice_Referral_to_Peripheral_Hospital=Count('id', filter=Q(referels__choice='Referral to Peripheral Hospital / Special Hospital for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Medical_College=Count('id', filter=Q(referels__choice='Referral to Medical College for management of Complication'), distinct=True),
                    Referral_choice_Referral_to_Private_facility=Count('id', filter=Q(referels__choice='Referral to Private facility'), distinct=True),
                    hypertension=Count('id', filter=Q(bloodPressure__gte=140), distinct=True),
                    total_LabTestAdded=Count('id', filter=Q(isLabTestAdded=True), distinct=True),
                    TestReportGenerated=Count('id', filter=Q(isLabTestReportGenerated=True), distinct=True)
                )

                # Aggregate counts for familySurvey_queryset
                familySurvey_data = familySurvey_queryset.aggregate(
                    total_family_count=Count('id', distinct=True),
                )

                combined_survey_data = {**healthpost_data, **familySurvey_data}

                Questionnaire_queryset = healthpost_queryset.filter(Questionnaire__isnull=False)
                suspected_disease_counts = get_suspected_disease_counts(Questionnaire_queryset)

                data_list.append({"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName,
                                "total_family_count":combined_survey_data["total_family_count"], "total_citizen_count":combined_survey_data["total_citizen_count"],
                                "total_cbac_count":combined_survey_data["total_cbac_count"], "citizen_above_60":combined_survey_data["citizen_above_60"],
                                "citizen_above_30":combined_survey_data["citizen_above_30"], "male":combined_survey_data["male"],
                                "female":combined_survey_data["female"], "transgender":combined_survey_data["transgender"], "total_AbhaCreated":combined_survey_data["total_AbhaCreated"],
                                "total_diabetes":suspected_disease_counts["diabetes"], "hypertension":combined_survey_data["hypertension"], "total_oral_cancer":suspected_disease_counts["oral_Cancer"],
                                "total_cervical_cancer":suspected_disease_counts["cervical_cancer"], "total_COPD_count":suspected_disease_counts["copd"],
                                "total_eye_problem":suspected_disease_counts["eye_disorder"], "total_ent_disorder":suspected_disease_counts["ent_disorder"],
                                "total_asthma":suspected_disease_counts["asthama"], "total_Alzheimers":suspected_disease_counts["Alzheimers"],
                                "total_tb_count":suspected_disease_counts["tb"], "total_leprosy":suspected_disease_counts["leprosy"],
                                "total_breast_cancer":suspected_disease_counts["breast_cancer"], "total_communicable":suspected_disease_counts["other_communicable_dieases"],
                                "blood_collected_home":combined_survey_data["blood_collected_home"], "blood_collected_center":combined_survey_data["blood_collected_center"],
                                "denieded_by_mo_count":combined_survey_data["denied_by_mo_count"] , "denieded_by_mo_individual":combined_survey_data["denied_by_mo_individual"],
                                "TestReportGenerated":combined_survey_data["TestReportGenerated"], "total_LabTestAdded":combined_survey_data["total_LabTestAdded"],
                                "Referral_choice_Referral_to_Mun_Dispensary":combined_survey_data["Referral_choice_Referral_to_Mun_Dispensary"],
                                "Referral_choice_Referral_to_HBT_polyclinic":combined_survey_data["Referral_choice_Referral_to_HBT_polyclinic"],
                                "Referral_choice_Referral_to_Peripheral_Hospital":combined_survey_data["Referral_choice_Referral_to_Peripheral_Hospital"],
                                "Referral_choice_Referral_to_Medical_College":combined_survey_data["Referral_choice_Referral_to_Medical_College"],
                                "Referral_choice_Referral_to_Private_facility":combined_survey_data["Referral_choice_Referral_to_Private_facility"],
                                "total_vulnerabel":combined_survey_data["total_vulnerable"]})

            page = self.paginate_queryset(data_list)
            if page is not None:
                return self.get_paginated_response(page)

            return Response({
                    "status":"success",
                    "message":"Successfully Fetched",
                    "data":data_list}, status = 200)

class GetAllUserDetails(generics.GenericAPIView):
    def get(self , request ):
        data = {'wards': []}

        wards = ward.objects.all().order_by("wardName")

        for w in wards:
            wrd = {'ward': w.wardName, 'healthPosts': []}

            healthposts = healthPost.objects.filter(ward=w)
            for hp in healthposts:
                areaData = area.objects.filter(healthPost_id = hp.id).values("areas")
                areaList =  [item["areas"] for item in areaData]
                health_post_info = {'healthPost': hp.healthPostName,'areaList':areaList, 'sections': []}

                sections = section.objects.filter(healthPost=hp)
                for sec in sections:
                    section_info = {'sectionName': sec.sectionName, 'anms': [],}

                    # ANMs
                    # anms = CustomUser.objects.filter(section_id=sec.id, groups__name="healthworker")
                    anms = CustomUser.objects.filter(userSections__in=[sec], groups__name="healthworker")

                    for anm in anms:
                        anm_info = {'anmName': anm.name, 'chvs': []}

                        # CHVs under the ANM
                        chvs = CustomUser.objects.filter(ANM_id = anm.id,groups__name="CHV-ASHA")
                        for chv in chvs:
                            chv_info = {'chvName': chv.name}
                            anm_info['chvs'].append(chv_info)

                        section_info['anms'].append(anm_info)

                    health_post_info['sections'].append(section_info)

                wrd['healthPosts'].append(health_post_info)

            data['wards'].append(wrd)

        return Response({
            'status': 'success',
            'message': 'Successfully Fetched',
            'data': data,
        })